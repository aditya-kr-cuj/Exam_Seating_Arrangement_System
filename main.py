import os
import pandas as pd
import sys
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Setting working directory
script_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(script_dir)

# Reading data from input Excel
input_file = "input_data_tt.xlsx"
room_df = pd.read_excel(input_file, sheet_name="in_room_capacity")
roll_name_df = pd.read_excel(input_file, sheet_name="in_roll_name_mapping")
course_roll_df = pd.read_excel(input_file, sheet_name="in_course_roll_mapping")
timetable_df = pd.read_excel(input_file, sheet_name="in_timetable")

while True:
    try:
        buffer = int(input("Enter No. of buffer seats to leave in each room: ").strip())
        if buffer >= 0:
            break
        else:
            print(" Buffer must be a non-negative integer.")
    except ValueError:
        print(" Please enter a valid integer.")

while True:
    mode = input("Enter allocation mode (dense or sparse): ").strip().lower()
    if mode in ["dense", "sparse"]:
        break
    print(" Invalid input. Please enter 'dense' or 'sparse'.")

print(" Starting processing...")

room_df.columns = [col.strip() for col in room_df.columns]
roll_name_df.columns = [col.strip() for col in roll_name_df.columns]
course_roll_df.columns = [col.strip() for col in course_roll_df.columns]

room_df["buffer_capacity"] = room_df["Exam Capacity"] - buffer

room_capacities = dict(zip(room_df["Room No."], room_df["buffer_capacity"]))
room_blocks = dict(zip(room_df["Room No."], room_df["Block"]))  # <== Block info

overall_seating_records = []
all_seat_summary_records = []

for idx, row in timetable_df.iterrows():
    exam_date = row["Date"]
    date_str = exam_date.strftime("%d_%m_%Y")
    base_folder = os.path.join(script_dir, date_str)
    os.makedirs(base_folder, exist_ok=True)

    seat_summary = []

    for session in ["Morning", "Evening"]:
        course_list = row.get(session)
        if pd.isna(course_list):
            continue

        session_folder = os.path.join(base_folder, session.lower())
        os.makedirs(session_folder, exist_ok=True)

        courses = [c.strip() for c in str(course_list).split(";")]
        print(f"\n Date: {date_str}, Session: {session}, Courses: {courses}")

        all_students = pd.DataFrame()

        for course_code in courses:
            course_students = course_roll_df[course_roll_df["course_code"] == course_code]
            if course_students.empty:
                continue
            rolls = course_students["Roll"].unique()
            student_names = roll_name_df[roll_name_df["Roll"].isin(rolls)]

            temp_df = pd.DataFrame({"Roll": rolls})
            temp_df = temp_df.merge(student_names, on="Roll", how="left")
            temp_df["Course Code"] = course_code
            all_students = pd.concat([all_students, temp_df], ignore_index=True)

        if all_students.empty:
            print(f"⚠ No students found for this session.")
            continue

        allocated = []
        room_list = list(room_capacities.keys())

        if mode == "dense":
            student_idx = 0
            for room in room_list:
                capacity = room_capacities[room]
                room_students = all_students.iloc[student_idx:student_idx + capacity].copy()
                room_students["Room No."] = room
                allocated.append(room_students)
                student_idx += capacity
                if student_idx >= len(all_students):
                    break
        else:  # sparse
            allocated_students = []
            num_rooms = len(room_list)
            for i, (_, student) in enumerate(all_students.iterrows()):
                room = room_list[i % num_rooms]
                student = student.copy()
                student["Room No."] = room
                allocated_students.append(student)
            allocated = [pd.DataFrame(allocated_students)]

        final_alloc = pd.concat(allocated, ignore_index=True)
        final_alloc["Date"] = date_str
        final_alloc["Session"] = session

        overall_seating_records.append(final_alloc)

        for (course_code, room), group in final_alloc.groupby(["Course Code", "Room No."]):
            filename = f"{date_str}_{course_code}_{room}_{session.lower()}.xlsx"

            filepath = os.path.join(session_folder, filename)

            output_df = group[["Roll", "Name", "Course Code", "Room No."]].copy()
            output_df["Signature"] = "" 

            output_df.to_excel(filepath, index=False)

            wb = openpyxl.load_workbook(filepath)
            ws = wb.active

            for cell in ws[1]:
                cell.font = Font(bold=True)

            footer_start_row = ws.max_row + 2

            ta_names = ["TA1", "TA2", "TA3", "TA4", "TA5"]
            invigilator_names = ["Invigilator1", "Invigilator2", "Invigilator3", "Invigilator4", "Invigilator5"]

            for i, name in enumerate(ta_names):
                ws.cell(row=footer_start_row + i, column=1, value=name)

            for i, name in enumerate(invigilator_names):
                ws.cell(row=footer_start_row + len(ta_names) + i, column=1, value=name)

            footer_rows = range(footer_start_row, footer_start_row + len(ta_names) + len(invigilator_names))
            for r in footer_rows:
                for c in range(1, 6):
                    ws.cell(row=r, column=c).alignment = Alignment(horizontal="left")

            wb.save(filepath)

        for room in final_alloc["Room No."].unique():
            students_allotted = final_alloc[final_alloc["Room No."] == room].shape[0]
            total_capacity = room_capacities[room]
            block = room_blocks.get(room, "N/A")
            seat_summary.append({
                "Date": date_str,
                "Session": session,
                "Block": block,  
                "Room No.": room,
                "Capacity (after buffer)": total_capacity,
                "Students Allotted": students_allotted,
                "Seats Left": total_capacity - students_allotted
            })

        print(f"Files saved in: {session_folder}")

    # Saving date-wise seat summary
    summary_df = pd.DataFrame(seat_summary)
    all_seat_summary_records.extend(seat_summary)
    summary_path = os.path.join(base_folder, f"seats_left_{date_str}.xlsx")
    summary_df.to_excel(summary_path, index=False)
    print(f"Seat summary written to: {summary_path}")

# Saving overall seating
overall_df = pd.concat(overall_seating_records, ignore_index=True)
overall_df[["Date", "Session", "Roll", "Name", "Course Code", "Room No."]].to_excel("overall_seating.xlsx", index=False)

# Saving full seat summary
all_summary_df = pd.DataFrame(all_seat_summary_records)
all_summary_df.to_excel("seats_left.xlsx", index=False)

print("\n Everything is created successfully.")